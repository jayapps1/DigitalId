from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from .models import User, Notification, IDRequest, IDRequestApproval
from .forms import IDRequestForm, AdminOfficerRegistrationForm, ContactMessageForm
from django.utils import timezone
from django.db.models import Q
from .models import User, OfficerProfile
from django.db import transaction
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.template.loader import render_to_string
from django.http import HttpResponse
from payments.models import Payment
from payments.utils import calculate_payment
import openpyxl
from django.views import View
from django.contrib.auth import get_user_model
import re
from django.db import transaction, models
from password_reset.sms_service import send_sms
from django.urls import reverse









def Home(request):
    return render(request, "digital_id/index.html")
#admin register view




User = get_user_model()



@login_required
def about(request):
    """
    Renders the About Digital ID page for logged-in officers only.
    """
    # Check if user is an officer (assuming officers have a staffid)
    if not hasattr(request.user, 'staffid') or not request.user.staffid:
        # Redirect non-officers (admins, public, etc.) to their dashboard or home
        if request.user.is_staff or request.user.is_superuser:
            return redirect('admin_dash:home')
        else:
            return redirect('digital_id:login')

    return render(request, 'about.html')


# -------------------------
# Excel Officer Import View
# -------------------------
class OfficerExcelImportView(View):
    template_name = 'digital_id/import_officer.html'

    def get(self, request):
        return render(request, self.template_name)

    def _safe_str(self, value, upper=False):
        if value is None:
            return None
        value = str(value).strip()
        return value.upper() if upper else value

    def post(self, request):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Please select an Excel file to upload.")
            return redirect('digital_id:import_officer')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception:
            messages.error(request, "Invalid Excel file.")
            return redirect('digital_id:import_officer')

        expected_columns = [
            'staffid', 'service_number', 'firstname', 'lastname', 'gender',
            'role', 'region', 'district', 'rank', 'station', 'phone'
        ]

        header = [str(cell.value).strip() if cell.value else ""
                  for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        if header != expected_columns:
            messages.error(
                request,
                f"Incorrect Excel columns. Expected: {', '.join(expected_columns)}"
            )
            return redirect('digital_id:import_officer')

        created = 0
        updated = 0
        skipped = []
        broadcast_list = []

        gender_map = {
            "MALE": "M", "FEMALE": "F", "M": "M", "F": "F"
        }

        rank_map = {
            "NON-COMMISSIONED OFFICER": "NCO",
            "COMMISSIONED OFFICER": "CO",
            "NCO": "NCO",
            "CO": "CO"
        }

        role_map = {
            "OFFICER": "OFFICER",
            "ADMIN OFFICER": "ADMIN",
            "SYSTEM SUPER ADMIN": "SUPERADMIN"
        }

        region_map = {
            "GREATER ACCRA": "GREATER_ACC",
            "ASHANTI": "ASHANTI",
            "NORTHERN": "NORTHERN",
            "UPPER EAST": "UPPER_EAST",
            "UPPER WEST": "UPPER_WEST",
            "WESTERN": "WESTERN",
            "EASTERN": "EASTERN",
            "CENTRAL": "CENTRAL",
            "BRONG AHAFO": "BRONG_AHAFO",
            "SAVANNAH": "SAVANNAH",
            "OTI": "OTI",
            "AHAFO": "AHAFO",
            "WESTERN NORTH": "WESTERN_NORTH",
            "BONO EAST": "BONO_EAST",
            "NORTH EAST": "NORTH_EAST",
            "VOLTA": "VOLTA",
        }

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            try:
                (
                    staffid, service_number, firstname, lastname, gender,
                    role, region, district, rank, station, phone
                ) = row
            except ValueError:
                skipped.append((idx, "Incorrect number of columns"))
                continue

            staffid = self._safe_str(staffid, upper=True)
            service_number = self._safe_str(service_number, upper=True)
            phone = self._safe_str(phone)
            region_clean = self._safe_str(region, upper=True)

            if not staffid or not service_number:
                skipped.append((idx, "Missing staffid or service_number"))
                continue

            # -------------------------
            # Normalize and validate service number
            # -------------------------
            pattern = r'^(GF\d{5,6}[A-Z]|\d{7,8}[A-Z])$'
            if not re.fullmatch(pattern, service_number):
                skipped.append((
                    idx,
                    "Invalid service number format. Must be:\n"
                    "- 7–8 digits followed by a letter (e.g., 2025500J)\n"
                    "- OR start with GF, followed by 5–6 digits and a letter (e.g., GF200058D)"
                ))
                continue

            gender_code = gender_map.get(self._safe_str(gender, upper=True))
            if not gender_code:
                skipped.append((idx, f"Invalid gender: {gender}"))
                continue

            region_code = region_map.get(region_clean)
            if not region_code:
                skipped.append((idx, f"Invalid region: {region}"))
                continue

            try:
                with transaction.atomic():
                    user, is_created = User.objects.update_or_create(
                        staffid=staffid,
                        defaults={
                            "service_number": service_number,
                            "firstname": self._safe_str(firstname),
                            "lastname": self._safe_str(lastname),
                            "gender": gender_code,
                            "role": role_map.get(
                                self._safe_str(role, upper=True),
                                "OFFICER"
                            ),
                            "region": region_code,
                            "district": self._safe_str(district),
                            "phone": phone,
                        }
                    )

                    if is_created:
                        user.set_password("Officer@123")
                        user.save()
                        created += 1
                        if phone:
                            broadcast_list.append((phone, user))
                    else:
                        updated += 1

                    OfficerProfile.objects.update_or_create(
                        user=user,
                        defaults={
                            "rank": rank_map.get(
                                self._safe_str(rank, upper=True),
                                "NCO"
                            ),
                            "station": self._safe_str(station),
                        }
                    )

            except Exception as e:
                skipped.append((idx, str(e)))

        # -------------------------
        # Broadcast SMS AFTER import
        # -------------------------
        for phone, officer in broadcast_list:
            try:
                sms_message = (
                    f"Hello {officer.firstname},\n\n"
                    f"Your Digital ID account is ready.\n"
                    f"Staff ID: {officer.staffid}\n"
                    f"Temporary Password: Officer@123\n\n"
                    f"Please change your password immediately:\n"
                    f"{request.build_absolute_uri(reverse('digital_id:change_password'))}"
                )
                send_sms(phone, sms_message)
            except Exception as e:
                print(f"SMS failed for {phone}: {e}")

        # -------------------------
        # Final Message
        # -------------------------
        message = f"{created} officers created, {updated} officers updated."
        if skipped:
            message += f" {len(skipped)} rows skipped.\n"
            for row_num, reason in skipped:
                message += f"Row {row_num}: {reason}\n"

        messages.success(request, message)
        return redirect('digital_id:import_officer')

# -------------------------
# contact us view
# -------------------------

@login_required
def contact_us(request):
    user = request.user

    # Only roles allowed to send messages
    send_roles = ["OFFICER", "STATION_ADMIN", "REGIONAL_ADMIN"]

    # Redirect roles that cannot send messages
    if user.role not in send_roles:
        # SUPERADMIN redirected to admin dashboard
        return redirect("admin_dash:home")

    # Prefill profile info if available
    officer_profile = getattr(user, "profile", None)

    if request.method == "POST":
        form = ContactMessageForm(request.POST, request.FILES)
        if form.is_valid():
            msg = form.save(commit=False)
            msg.user = user
            msg.role = user.role  # make role immutable
            msg.save()
            # Redirect to success page after sending
            return redirect("digital_id:contact_success")
    else:
        form = ContactMessageForm()

    return render(
        request,
        "digital_id/contact_us.html",
        {
            "form": form,
            "officer_profile": officer_profile,  # pass profile for template prefill
        }
    )


# -------------------------
# message success view
# -------------------------

@login_required
def contact_success(request):
    return render(request, "digital_id/contact_success.html")



# -------------------------
# Officer ID / Print Page
# -------------------------
def officer_id(request, staffid):
    profile = get_object_or_404(OfficerProfile, user__staffid=staffid)
    return render(request, "digital_id/officer_id.html", {
        "profile": profile,
        "hide_qr": False,
        "is_verify_view": False,
    })




# -------------------------
# Admin check decorator
# -------------------------
def is_admin(user):
    return user.is_authenticated and user.role in [
        "SUPERADMIN",      # your top-level admin
        "REGIONAL_ADMIN", # regional admin
        "STATION_ADMIN",  # station admin
    ]



@login_required
@user_passes_test(is_admin)
def register_officer(request):
    if request.method == "POST":
        form = AdminOfficerRegistrationForm(request.POST, request=request)
        if form.is_valid():
            try:
                with transaction.atomic():
                    officer = form.save()  # Save the new officer
            except IntegrityError as e:
                form.add_error(None, f"Registration failed: Duplicate entry ({str(e)})")
            else:
                # -------- Send SMS using officer.phone (normalized) --------
                if officer.phone:
                    try:
                        sms_result = send_sms(
                            officer.phone,
                            f"Hello {officer.firstname},\n\n"
                            f"Your Digital ID account is ready.\n"
                            f"Staff ID: {officer.staffid}\n"
                            f"Temporary Password: Officer@123\n\n"
                            f"Please change your password immediately:\n"
                            f"{request.build_absolute_uri(reverse('digital_id:change_password'))}"
                        )
                        if not sms_result.get("success"):
                            messages.warning(
                                request,
                                f"Officer registered, but SMS failed: {sms_result.get('error')}"
                            )
                    except Exception as e:
                        messages.warning(
                            request,
                            f"Officer registered, but SMS sending encountered an error: {str(e)}"
                        )

                messages.success(
                    request,
                    "Officer registered successfully. Default password assigned if none provided."
                )
                return redirect("digital_id:register_officer")
    else:
        form = AdminOfficerRegistrationForm(request=request)

    return render(
        request,
        "digital_id/register_officer.html",
        {"form": form}
    )


@login_required
def request_id_view(request):
    user = request.user
    profile = user.profile

    if not user.profile_completed:
        messages.error(request, "You must complete your profile before requesting a new ID card.")
        return redirect("officers_dash:complete_profile")

    unpaid_request = IDRequest.objects.filter(
        officer=user
    ).filter(
        Q(payment__status__in=["PENDING", "FAILED"]) | Q(payment__isnull=True)
    ).order_by("-date_requested").first()

    if unpaid_request:
        messages.warning(request, "You have a pending payment. Please complete it before requesting a new ID.")
        return redirect("payments_sys:resume_payment", staffid=user.staffid)

    has_new = IDRequest.objects.filter(officer=user, request_type="NEW").exists()

    if not has_new:
        allowed_types = ["NEW"]
    else:
        allowed_types = ["LOST"]
        if profile.qr_expiry_date and profile.qr_expiry_date <= timezone.now():
            allowed_types = ["EXPIRED"]

    if request.method == "POST":
        form = IDRequestForm(request.POST, user=user)
        if form.is_valid():
            request_type = form.cleaned_data["request_type"]
            if request_type not in allowed_types:
                messages.error(request, "Invalid request type selection.")
                return redirect("digital_id:request_id")

            # 🔹 Create new IDRequest and Approval immediately
            id_request = IDRequest.objects.create(
                officer=user,
                request_type=request_type
            )
            IDRequestApproval.objects.get_or_create(id_request=id_request)

            # Deactivate QR until approval
            profile.is_active_qr = False
            profile.save(update_fields=["is_active_qr"])

            # Redirect to payment
            return redirect("payments_sys:start_new_id_payment", request_type=id_request.request_type)
    else:
        form = IDRequestForm(user=user)

    request_list = IDRequest.objects.filter(
        officer=user
    ).select_related("payment", "approval").order_by("-date_requested")

    has_pending_request = request_list.filter(
        payment__status="SUCCESS",
        approval__status="PENDING"
    ).exists()

    has_paid = request_list.filter(payment__status="SUCCESS").exists()

    return render(request, "digital_id/idrequest_form.html", {
        "form": form,
        "allowed_types": allowed_types,
        "id_requests": request_list,
        "unpaid_request": unpaid_request,
        "has_pending_request": has_pending_request,
        "has_paid": has_paid,
        "profile": profile,
    })



def user_login(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f"Welcome, {user.firstname}")

            # Force officers to change default password
            if user.is_officer and user.must_change_password:
                Notification.objects.create(
                    user=user,
                    message="You must change your default password before proceeding."
                )
                return redirect("digital_id:change_password")

            # Role-based redirects
            if user.is_superuser:
                return redirect("/admin/admin-dash")
            elif user.is_staff or user.is_officer:
                return redirect("officers_dash:dashboard")
            else:
                messages.error(request, "Access denied.")
                logout(request)
                return redirect("digital_id:login")
        else:
            messages.error(request, "Invalid credentials.")
    else:
        form = AuthenticationForm()

    return render(request, "digital_id/login.html", {"form": form})



@login_required
def change_password(request):
    user = request.user

    # Determine if this is an officer forced to change password
    is_officer = getattr(user, "is_officer", False)

    if request.method == "POST":
        form = PasswordChangeForm(user, request.POST)
        if form.is_valid():
            form.save()
            # Only reset must_change_password for officers
            if is_officer:
                user.must_change_password = False
                user.save(update_fields=["must_change_password"])

            # Keep the user logged in
            update_session_auth_hash(request, user)

            # Optional: create notification for officer only
            if is_officer:
                Notification.objects.create(
                    user=user,
                    message="Your password has been successfully updated."
                )

            messages.success(request, "Password updated successfully.")

            # Redirect based on profile completion (officers only)
            if is_officer and not user.profile_completed:
                return redirect("officers_dash:complete_profile")

            # Redirect to dashboard based on user type
            if user.is_staff or user.is_superuser:
                return redirect("admin_dash:home")
            else:
                return redirect("officers_dash:dashboard")

        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = PasswordChangeForm(user)

    return render(request, "digital_id/change_password.html", {"form": form})




from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView


def user_logout(request):
    logout(request)
    messages.success(request, "You have successfully logged out.")
    return redirect('digital_id:login')



class NotificationsView(LoginRequiredMixin, TemplateView):
    template_name = "digital_id/notifications.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # SUPERADMIN → sees all
        if user.role == "SUPERADMIN":
            notifications = Notification.objects.select_related("user")

        # REGIONAL ADMIN → only their region
        elif user.role == "REGIONAL_ADMIN":
            notifications = Notification.objects.select_related("user").filter(
                user__region=user.region
            )

        # STATION ADMIN → only their station
        elif user.role == "STATION_ADMIN":
            notifications = Notification.objects.select_related("user").filter(
                user__profile__station=user.profile.station
            )

        # OFFICER → only their own
        else:
            notifications = Notification.objects.filter(user=user)

        # Mark only what the user can see as read
        notifications.filter(is_read=False).update(is_read=True)

        context["notifications"] = notifications.order_by("-created_at")
        return context



from django.views.decorators.http import require_POST
@login_required
def view_notification(request, pk):
    user = request.user

    if user.role == "SUPERADMIN":
        queryset = Notification.objects.all()

    elif user.role == "REGIONAL_ADMIN":
        queryset = Notification.objects.filter(user__region=user.region)

    elif user.role == "STATION_ADMIN":
        queryset = Notification.objects.filter(
            user__profile__station=user.profile.station
        )

    else:
        queryset = Notification.objects.filter(user=user)

    note = queryset.filter(pk=pk).first()

    # 🚨 Graceful failure instead of 404
    if not note:
        messages.warning(
            request,
            "You are not permitted to view this notification."
        )
        return redirect("digital_id:notifications")

    # Mark as read
    if not note.is_read:
        note.is_read = True
        note.save(update_fields=["is_read"])

    # Redirect if notification has a link
    if note.link:
        return redirect(note.link)

    return render(
        request,
        "digital_id/notification_detail.html",
        {"note": note}
    )



@login_required
def delete_notification(request, pk):
    user = request.user

    qs = Notification.objects.filter(pk=pk)

    if user.role == "SUPERADMIN":
        pass
    elif user.role == "REGIONAL_ADMIN":
        qs = qs.filter(user__region=user.region)
    elif user.role == "STATION_ADMIN":
        qs = qs.filter(user__profile__station=user.profile.station)
    else:
        qs = qs.filter(user=user)

    qs.delete()
    return redirect("digital_id:notifications")


@login_required
@require_POST
def bulk_delete_notifications(request):
    ids = request.POST.getlist("notification_ids")
    user = request.user

    qs = Notification.objects.filter(id__in=ids)

    if user.role == "SUPERADMIN":
        pass
    elif user.role == "REGIONAL_ADMIN":
        qs = qs.filter(user__region=user.region)
    elif user.role == "STATION_ADMIN":
        qs = qs.filter(user__profile__station=user.profile.station)
    else:
        qs = qs.filter(user=user)

    qs.delete()
    return redirect("digital_id:notifications")



@login_required
def mark_as_read(request, pk):
    """Mark a single notification as read via AJAX."""
    note = Notification.objects.filter(pk=pk, user=request.user).first()
    if note and not note.is_read:
        note.is_read = True
        note.save(update_fields=['is_read'])
    return JsonResponse({'success': True})

@login_required
def get_unread_count(request):
    """Return unread notifications count via AJAX."""
    count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'unread_count': count})
